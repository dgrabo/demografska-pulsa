#!/usr/bin/env python3
"""
Process DZS migration data and generate public/data/migracije.json.

Data sources:
  - stanovnistvo-migracije.xlsx  (sheet 7.2.1.: 2001-2024 by country)
  - stan-2025-2-1_tablice-hr.xlsx (I T5: county external 2020-2024,
                                    II T3: internal 2024,
                                    III T1: combined 2024)
  - stan-2024-2-1_tablice_hr.xlsx (I T5: county external 2019-2023)
  - stan-2023-2-1_tablice_hr.xlsx (I T5: county external 2018-2022)
  - stan-2022-2-1_tablice-hr.xlsx (I T5: county external 2017-2021)
  - stan-2021-2-1_tablice-hr.xlsx (I T5: county external 2016-2020)
"""

import json
import os
import openpyxl

RAW_DIR = os.path.join(os.path.dirname(__file__), 'raw')
OUT_PATH = os.path.join(os.path.dirname(__file__), '..', 'public', 'data', 'migracije.json')

# County name -> ID mapping (reused from process_usluge.py)
COUNTY_NAME_TO_ID = {
    'Zagrebacka': 'ZK',
    'Krapinsko-zagorska': 'KR',
    'Sisacko-moslavacka': 'SM',
    'Karlovacka': 'KA',
    'Varazdinska': 'VZ',
    'Koprivnicko-krizevacka': 'KZ',
    'Bjelovarsko-bilogorska': 'BB',
    'Primorsko-goranska': 'PG',
    'Licko-senjska': 'LS',
    'Viroviticko-podravska': 'VP',
    'Pozesko-slavonska': 'PS',
    'Pozeko-slavonska': 'PS',
    'Brodsko-posavska': 'BPZ',
    'Zadarska': 'ZD',
    'Osjecko-baranjska': 'OB',
    'Sibensko-kninska': 'SK',
    'Vukovarsko-srijemska': 'VK',
    'Splitsko-dalmatinska': 'SD',
    'Istarska': 'IS',
    'Dubrovacko-neretvanska': 'DN',
    'Medimurska': 'ME',
    'Grad Zagreb': 'ZG',
}

COUNTY_ID_TO_NAME = {
    'ZK': 'Zagrebačka',
    'KR': 'Krapinsko-zagorska',
    'SM': 'Sisačko-moslavačka',
    'KA': 'Karlovačka',
    'VZ': 'Varaždinska',
    'KZ': 'Koprivničko-križevačka',
    'BB': 'Bjelovarsko-bilogorska',
    'PG': 'Primorsko-goranska',
    'LS': 'Ličko-senjska',
    'VP': 'Virovitičko-podravska',
    'PS': 'Požeško-slavonska',
    'BPZ': 'Brodsko-posavska',
    'ZD': 'Zadarska',
    'OB': 'Osječko-baranjska',
    'SK': 'Šibensko-kninska',
    'VK': 'Vukovarsko-srijemska',
    'SD': 'Splitsko-dalmatinska',
    'IS': 'Istarska',
    'DN': 'Dubrovačko-neretvanska',
    'ME': 'Međimurska',
    'ZG': 'Grad Zagreb',
}


def clean_int(val):
    """Convert a cell value to int, handling dashes, non-breaking spaces, footnotes."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    s = str(val).strip()
    if s in ('-', '', '…', '...'):
        return 0
    # Remove footnote markers like "1)" at the end
    import re
    s = re.sub(r'\d+\)$', '', s).strip()
    # Remove all spaces and non-breaking spaces
    s = s.replace('\xa0', '').replace(' ', '').replace('\u200b', '')
    if s in ('-', '', '…', '...'):
        return 0
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return 0


def normalize_county_name(name):
    """Normalize a county name string for lookup."""
    if name is None:
        return ''
    s = str(name).strip()
    # Remove diacritics for matching
    replacements = {
        'č': 'c', 'ć': 'c', 'đ': 'd', 'š': 's', 'ž': 'z',
        'Č': 'C', 'Ć': 'C', 'Đ': 'D', 'Š': 'S', 'Ž': 'Z',
    }
    for src, dst in replacements.items():
        s = s.replace(src, dst)
    return s


def county_name_to_id(name):
    """Map a county name from Excel to its two/three-letter ID."""
    norm = normalize_county_name(name)
    if norm in COUNTY_NAME_TO_ID:
        return COUNTY_NAME_TO_ID[norm]
    # Try without trailing " zupanija"
    for suffix in [' zupanija', ' Zupanija', ' županija', ' Županija']:
        clean = name.strip().rstrip('.')
        if clean.endswith(suffix):
            clean = clean[:-len(suffix)].strip()
            norm2 = normalize_county_name(clean)
            if norm2 in COUNTY_NAME_TO_ID:
                return COUNTY_NAME_TO_ID[norm2]
    return None


# ─────────────────────────────────────────────────
# A) National time series + country breakdowns from stanovnistvo-migracije.xlsx
# ─────────────────────────────────────────────────

def read_national_migration():
    """Read sheet 7.2.1. for 2001-2024 national data + per-country breakdown."""
    fpath = os.path.join(RAW_DIR, 'stanovnistvo-migracije.xlsx')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb['7.2.1.']

    # Parse year columns from row 6
    # Years are in odd columns (C, E, G, ...) starting at col 3
    # Each year has two sub-columns: doseljeni (odd col), odseljeni (even col)
    year_cols = {}  # year -> (doseljeni_col, odseljeni_col)
    for col_idx in range(3, ws.max_column + 1):
        val = ws.cell(row=6, column=col_idx).value
        if val is not None:
            year_str = str(val).strip().rstrip('.')
            try:
                year = int(year_str)
                if 2000 <= year <= 2030:
                    year_cols[year] = (col_idx, col_idx + 1)
            except ValueError:
                pass

    years = sorted(year_cols.keys())
    print(f"  Years found in 7.2.1.: {years[0]}-{years[-1]} ({len(years)} years)")

    # Row mapping for countries of interest
    # Read all rows to build a name->row mapping
    country_rows = {}
    for row_idx in range(9, 60):
        name = ws.cell(row=row_idx, column=1).value
        if name:
            # Normalize whitespace: collapse multiple spaces to single
            import re as _re
            cleaned = _re.sub(r'\s+', ' ', str(name).strip())
            country_rows[cleaned] = row_idx

    # National totals (row 9 = "Ukupno")
    nacional = []
    for year in years:
        d_col, o_col = year_cols[year]
        doseljeni = clean_int(ws.cell(row=9, column=d_col).value)
        odseljeni = clean_int(ws.cell(row=9, column=o_col).value)
        nacional.append({
            'godina': year,
            'doseljeni': doseljeni,
            'odseljeni': odseljeni,
            'saldo': doseljeni - odseljeni,
        })

    # Emigration destinations (odseljeni column)
    emigration_countries = {
        'Njemačka': 'Njemacka',
        'Austrija': 'Austrija',
        'Irska': 'Irska',
        'Švicarska': 'Svicarska',
        'Švedska': 'Svedska',
        'Slovenija': 'Slovenija',
    }

    drzave_emigracija = {}
    for display_name, lookup_name in emigration_countries.items():
        row_idx = country_rows.get(lookup_name)
        if not row_idx:
            # Try with diacritics
            for k, v in country_rows.items():
                if normalize_county_name(k) == lookup_name:
                    row_idx = v
                    break
        if not row_idx:
            print(f"  WARNING: Could not find row for emigration country '{lookup_name}'")
            continue
        series = []
        for year in years:
            d_col, o_col = year_cols[year]
            val = clean_int(ws.cell(row=row_idx, column=o_col).value)
            series.append({'godina': year, 'broj': val})
        drzave_emigracija[display_name] = series

    # Compute "Ostale" for emigration = total odseljeni - sum of tracked countries
    ostale_emig = []
    for i, year in enumerate(years):
        total_odseljeni = nacional[i]['odseljeni']
        tracked = sum(
            drzave_emigracija[c][i]['broj']
            for c in drzave_emigracija
        )
        ostale_emig.append({'godina': year, 'broj': max(0, total_odseljeni - tracked)})
    drzave_emigracija['Ostale'] = ostale_emig

    # Immigration origins (doseljeni column)
    immigration_countries = {
        'Bosna i Hercegovina': 'Bosna i Hercegovina',
        'Srbija': 'Srbija',
        'Slovenija (im.)': None,  # skip — already in emigration
        'Kosovo': None,
        'Sjeverna Makedonija': None,
        'Azija': 'Azija',
    }

    # Build immigration from specific rows
    imm_targets = {
        'Bosna i Hercegovina': None,
        'Srbija': None,
        'Azija': None,
        'Sjeverna Makedonija': None,
    }

    for k in country_rows:
        # Strip footnote markers (e.g., "3)") and normalize
        import re as _re2
        clean_k = _re2.sub(r'\d+\)\s*$', '', k).strip()
        norm = normalize_county_name(clean_k)
        if norm == 'Bosna i Hercegovina':
            imm_targets['Bosna i Hercegovina'] = country_rows[k]
        elif norm == 'Srbija':
            imm_targets['Srbija'] = country_rows[k]
        elif norm == 'Azija':
            imm_targets['Azija'] = country_rows[k]
        elif norm == 'Sjeverna Makedonija':
            imm_targets['Sjeverna Makedonija'] = country_rows[k]

    drzave_imigracija = {}
    for display_name, row_idx in imm_targets.items():
        if row_idx is None:
            print(f"  WARNING: Could not find row for immigration country '{display_name}'")
            continue
        series = []
        for year in years:
            d_col, o_col = year_cols[year]
            val = clean_int(ws.cell(row=row_idx, column=d_col).value)
            series.append({'godina': year, 'broj': val})
        drzave_imigracija[display_name] = series

    # Compute "Ostale" for immigration
    ostale_imm = []
    for i, year in enumerate(years):
        total_doseljeni = nacional[i]['doseljeni']
        tracked = sum(
            drzave_imigracija[c][i]['broj']
            for c in drzave_imigracija
        )
        ostale_imm.append({'godina': year, 'broj': max(0, total_doseljeni - tracked)})
    drzave_imigracija['Ostale'] = ostale_imm

    wb.close()
    return nacional, drzave_emigracija, drzave_imigracija


# ─────────────────────────────────────────────────
# B) County-level external migration from I T5 sheets (2016-2024)
# ─────────────────────────────────────────────────

def read_county_external(filename, sheet_name='I T5'):
    """Read county-level doseljeni/odseljeni from an annual DZS file's I T5 sheet."""
    fpath = os.path.join(RAW_DIR, filename)
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb[sheet_name]

    # Parse year columns — scan rows 3 and 4 (different files use different rows)
    year_cols = {}
    for header_row in (3, 4):
        for col_idx in range(2, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col_idx).value
            if val is not None:
                year_str = str(val).strip().rstrip('.')
                try:
                    year = int(year_str)
                    if 2010 <= year <= 2030:
                        year_cols[year] = (col_idx, col_idx + 1)
                except ValueError:
                    pass

    # Find first county data row by scanning for "Republika Hrvatska" or county names
    data_start = 5
    for row_idx in range(4, 10):
        name = ws.cell(row=row_idx, column=1).value
        if name and 'Republika Hrvatska' in str(name):
            data_start = row_idx + 1
            break

    # Read county rows
    result = {}  # county_id -> { year: { doseljeni, odseljeni } }
    for row_idx in range(data_start, data_start + 25):
        name = ws.cell(row=row_idx, column=1).value
        if name is None:
            continue
        cid = county_name_to_id(name)
        if cid is None:
            continue
        if cid not in result:
            result[cid] = {}
        for year, (d_col, o_col) in year_cols.items():
            doseljeni = clean_int(ws.cell(row=row_idx, column=d_col).value)
            odseljeni = clean_int(ws.cell(row=row_idx, column=o_col).value)
            result[cid][year] = {'doseljeni': doseljeni, 'odseljeni': odseljeni}

    wb.close()
    return result, sorted(year_cols.keys())


def merge_county_external():
    """Merge I T5 sheets across all annual files to get 2016-2024."""
    files = [
        ('stan-2021-2-1_tablice-hr.xlsx', 'I T5'),   # 2016-2020
        ('stan-2022-2-1_tablice-hr.xlsx', 'I T5'),    # 2017-2021
        ('stan-2023-2-1_tablice_hr.xlsx', 'I T5'),    # 2018-2022
        ('stan-2024-2-1_tablice_hr.xlsx', 'I T5'),    # 2019-2023
        ('stan-2025-2-1_tablice-hr.xlsx', 'I T5'),    # 2020-2024
    ]

    merged = {}  # county_id -> { year -> { doseljeni, odseljeni } }
    all_years = set()

    for filename, sheet in files:
        fpath = os.path.join(RAW_DIR, filename)
        if not os.path.exists(fpath):
            print(f"  Skipping {filename} (not found)")
            continue
        print(f"  Reading {filename} {sheet}...")
        data, years = read_county_external(filename, sheet)
        all_years.update(years)
        for cid, year_data in data.items():
            if cid not in merged:
                merged[cid] = {}
            for year, vals in year_data.items():
                # Later files overwrite earlier ones for overlapping years (more recent data)
                merged[cid][year] = vals

    return merged, sorted(all_years)


# ─────────────────────────────────────────────────
# C) Internal migration saldo from II T3
# ─────────────────────────────────────────────────

def read_internal_migration():
    """Read inter-county migration saldo from II T3 of the latest annual file."""
    fpath = os.path.join(RAW_DIR, 'stan-2025-2-1_tablice-hr.xlsx')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb['II T3']

    # Column F = "Saldo migracije medu zupanijama" (net inter-county)
    result = {}
    for row_idx in range(7, 30):
        name = ws.cell(row=row_idx, column=1).value
        if name is None:
            continue
        cid = county_name_to_id(name)
        if cid is None:
            continue
        saldo = clean_int(ws.cell(row=row_idx, column=6).value)
        result[cid] = saldo

    wb.close()
    return result


# ─────────────────────────────────────────────────
# D) Combined total migration from III T1
# ─────────────────────────────────────────────────

def read_combined_migration():
    """Read combined (external + internal) saldo from III T1."""
    fpath = os.path.join(RAW_DIR, 'stan-2025-2-1_tablice-hr.xlsx')
    wb = openpyxl.load_workbook(fpath, data_only=True)
    ws = wb['III T1']

    # Column H = "Saldo ukupne migracije"
    # Column I = "Saldo migracije medu zupanijama"
    # Column J = "Saldo migracije s inozemstvom"
    result = {}
    for row_idx in range(7, 30):
        name = ws.cell(row=row_idx, column=1).value
        if name is None:
            continue
        cid = county_name_to_id(name)
        if cid is None:
            continue

        saldo_ukupni = clean_int(ws.cell(row=row_idx, column=8).value)
        saldo_medu_zup = clean_int(ws.cell(row=row_idx, column=9).value)
        saldo_inozemstvo = clean_int(ws.cell(row=row_idx, column=10).value)

        result[cid] = {
            'saldo_ukupni': saldo_ukupni,
            'saldo_medu_zupanijama': saldo_medu_zup,
            'saldo_inozemstvo': saldo_inozemstvo,
        }

    wb.close()
    return result


# ─────────────────────────────────────────────────
# Main: assemble and export
# ─────────────────────────────────────────────────

def main():
    print("=== Processing migration data ===\n")

    # A) National trend + country breakdowns
    print("1. Reading national migration data (7.2.1.)...")
    nacional, drzave_emig, drzave_imig = read_national_migration()
    print(f"   National trend: {len(nacional)} years")
    print(f"   Emigration countries: {list(drzave_emig.keys())}")
    print(f"   Immigration countries: {list(drzave_imig.keys())}")

    # Summary stats
    total_emig_since_2013 = sum(
        n['odseljeni'] for n in nacional if n['godina'] >= 2013
    )
    total_imig_since_2013 = sum(
        n['doseljeni'] for n in nacional if n['godina'] >= 2013
    )
    latest = nacional[-1]
    print(f"   Since 2013: {total_emig_since_2013:,} emigrated, {total_imig_since_2013:,} immigrated")
    print(f"   Latest year ({latest['godina']}): saldo = {latest['saldo']:+,}")

    # B) County external migration (2016-2024)
    print("\n2. Reading county-level external migration (I T5)...")
    county_ext, ext_years = merge_county_external()
    print(f"   Counties: {len(county_ext)}, Years: {ext_years}")

    # C) Internal migration saldo
    print("\n3. Reading internal migration saldo (II T3)...")
    internal = read_internal_migration()
    print(f"   Counties with internal saldo: {len(internal)}")

    # D) Combined migration
    print("\n4. Reading combined migration (III T1)...")
    combined = read_combined_migration()
    print(f"   Counties with combined data: {len(combined)}")

    # Build po_zupanijama
    po_zupanijama = {}
    for cid in sorted(set(list(county_ext.keys()) + list(internal.keys()) + list(combined.keys()))):
        naziv = COUNTY_ID_TO_NAME.get(cid, cid)

        # External migration time series
        vanjska = []
        if cid in county_ext:
            for year in sorted(county_ext[cid].keys()):
                d = county_ext[cid][year]
                vanjska.append({
                    'godina': year,
                    'doseljeni': d['doseljeni'],
                    'odseljeni': d['odseljeni'],
                    'saldo': d['doseljeni'] - d['odseljeni'],
                })

        # Internal migration saldo (latest year only)
        unutarnja_saldo = internal.get(cid, 0)

        # Combined saldo
        comb = combined.get(cid, {})

        po_zupanijama[cid] = {
            'naziv': naziv,
            'vanjska': vanjska,
            'unutarnja_saldo': unutarnja_saldo,
            'ukupni_saldo_zadnja_godina': comb.get('saldo_ukupni', 0),
            'saldo_inozemstvo_zadnja': comb.get('saldo_inozemstvo', 0),
        }

    # Assemble output
    output = {
        'nacionalni_trend': nacional,
        'drzave_emigracija': drzave_emig,
        'drzave_imigracija': drzave_imig,
        'po_zupanijama': po_zupanijama,
    }

    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"\n=== Output written to {OUT_PATH} ===")
    print(f"    File size: {os.path.getsize(OUT_PATH) / 1024:.1f} KB")

    # Print summary table
    print("\n--- County migration summary (latest year) ---")
    print(f"{'County':<30} {'Ext. saldo':>12} {'Int. saldo':>12} {'Total':>12}")
    print('-' * 70)
    for cid in sorted(po_zupanijama.keys(), key=lambda x: po_zupanijama[x].get('ukupni_saldo_zadnja_godina', 0)):
        z = po_zupanijama[cid]
        ext_s = z['vanjska'][-1]['saldo'] if z['vanjska'] else 0
        print(f"{z['naziv']:<30} {ext_s:>+12,} {z['unutarnja_saldo']:>+12,} {z['ukupni_saldo_zadnja_godina']:>+12,}")


if __name__ == '__main__':
    main()
