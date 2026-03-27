#!/usr/bin/env python3
"""process_usluge.py — Process school enrollment and GP availability data by county.

Data sources:
- obr-2025-2-1_tablice-hrv.xlsx (T2.2) — 2024/2025 primary school pupil counts
- skole_2010.xlsx (Table 6) — 2010/2011 primary school pupil counts
- web_opca_022026.xlsx (28.02.2026) — GP surgery data with patient counts

Outputs:
- ../public/data/usluge.json — county-level school + healthcare + risk data
"""

import json
import os
import openpyxl

RAW_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'raw')
OUT_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'public', 'data', 'usluge.json')
ZUPANIJE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'public', 'data', 'zupanije.json')

# County name → ID mapping (matching zupanije.json)
COUNTY_NAME_TO_ID = {
    'Zagrebačka': 'ZK',
    'Krapinsko-zagorska': 'KR',
    'Sisačko-moslavačka': 'SM',
    'Karlovačka': 'KA',
    'Varaždinska': 'VZ',
    'Koprivničko-križevačka': 'KZ',
    'Bjelovarsko-bilogorska': 'BB',
    'Primorsko-goranska': 'PG',
    'Ličko-senjska': 'LS',
    'Virovitičko-podravska': 'VP',
    'Požeško-slavonska': 'PS',
    'Brodsko-posavska': 'BPZ',
    'Zadarska': 'ZD',
    'Osječko-baranjska': 'OB',
    'Šibensko-kninska': 'SK',
    'Vukovarsko-srijemska': 'VK',
    'Splitsko-dalmatinska': 'SD',
    'Istarska': 'IS',
    'Dubrovačko-neretvanska': 'DN',
    'Međimurska': 'ME',
    'Grad Zagreb': 'ZG',
}

# 2010 file uses "županija" suffix
COUNTY_NAME_2010_TO_ID = {
    'Karlovačka županija': 'KA',
    'Varaždinska županija': 'VZ',
    'Koprivničko-križevačka županija': 'KZ',
    'Bjelovarsko-bilogorska županija': 'BB',
    'Primorsko-goranska županija': 'PG',
    'Ličko-senjska županija': 'LS',
    'Virovitičko-podravska županija': 'VP',
    'Požeško-slavonska županija': 'PS',
    'Brodsko-posavska županija': 'BPZ',
    'Zadarska županija': 'ZD',
    'Osječko-baranjska županija': 'OB',
    'Šibensko-kninska županija': 'SK',
    'Vukovarsko-srijemska županija': 'VK',
    'Splitsko-dalmatinska županija': 'SD',
    'Istarska županija': 'IS',
    'Dubrovačko-neretvanska županija': 'DN',
    'Međimurska županija': 'ME',
    'Grad Zagreb': 'ZG',
}

# PU → County ID mapping (HZZO regional offices → counties)
PU_TO_COUNTY = {
    'PU Bjelovar': 'BB',
    'PU Dubrovnik': 'DN',
    'PU Gospić': 'LS',
    'PU Karlovac': 'KA',
    'PU Koprivnica': 'KZ',
    'PU Krapina': 'KR',
    'PU Osijek': 'OB',
    'PU Pazin': 'IS',
    'PU Požega': 'PS',
    'PU Rijeka': 'PG',
    'PU Sisak': 'SM',
    'PU Slavonski Brod': 'BPZ',
    'PU Split': 'SD',
    'PU Varaždin': 'VZ',
    'PU Vinkovci': 'VK',
    'PU Virovitica': 'VP',
    'PU Zadar': 'ZD',
    'PU Čakovec': 'ME',
    'PU Šibenik': 'SK',
}

# Zagreb city districts (Grad Zagreb = ZG), everything else in PU Zagreb → ZK
ZAGREB_DISTRICTS = {
    'DONJI GRAD', 'GORNJI GRAD - MEDVEŠČAK', 'TRNJE', 'MAKSIMIR',
    'PEŠČENICA-ŽITNJAK', 'NOVI ZAGREB - ISTOK', 'NOVI ZAGREB - ZAPAD',
    'TREŠNJEVKA - SJEVER', 'TREŠNJEVKA - JUG', 'TREŠNJEVKA-SJEVER',
    'TREŠNJEVKA-JUG', 'ČRNOMEREC', 'GORNJA DUBRAVA', 'DONJA DUBRAVA',
    'STENJEVEC', 'PODSUSED-VRAPČE', 'PODSLJEME', 'SESVETE',
    'BREZOVICA', 'GRAD ZAGREB', 'DUBRAVA',
}


def read_school_2024():
    """Read 2024/2025 primary school pupil counts from T2.2 sheet."""
    path = os.path.join(RAW_DIR, 'obr-2025-2-1_tablice-hrv.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['T2.2']

    school_data = {}
    # Rows 7-27 contain individual counties (row 6 is national total)
    for row in ws.iter_rows(min_row=7, max_row=27, values_only=True):
        name = row[0]
        pupils = row[4]
        if name and pupils:
            name = name.strip()
            county_id = COUNTY_NAME_TO_ID.get(name)
            if county_id:
                school_data[county_id] = int(pupils)
            else:
                print(f"  Warning: Could not map county name '{name}' (2024)")

    print(f"School 2024: {len(school_data)} counties, total {sum(school_data.values()):,} pupils")
    return school_data


def read_school_2010():
    """Read 2010/2011 primary school pupil counts from Table 6 sheet."""
    path = os.path.join(RAW_DIR, 'skole_2010.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['Table 6']

    school_data = {}
    # Rows 1-18: county data, column J (index 9) = total pupils
    for row in ws.iter_rows(min_row=1, max_row=18, values_only=True):
        name = row[0]
        pupils = row[9]
        if name and pupils:
            name = name.strip()
            county_id = COUNTY_NAME_2010_TO_ID.get(name)
            if county_id:
                school_data[county_id] = int(pupils)
            else:
                print(f"  Warning: Could not map county name '{name}' (2010)")

    # Hardcode 3 missing counties
    school_data['ZK'] = 27314   # Zagrebačka
    school_data['KR'] = 11160   # Krapinsko-zagorska
    school_data['SM'] = 14172   # Sisačko-moslavačka

    print(f"School 2010: {len(school_data)} counties, total {sum(school_data.values()):,} pupils")
    return school_data


def read_gp_data():
    """Read GP surgery data, aggregate by county."""
    path = os.path.join(RAW_DIR, 'web_opca_022026.xlsx')
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb['28.02.2026']

    # Per-county aggregation: {county_id: {doctors: int, patients: int}}
    gp_data = {}

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        pu_name = row[0]
        patients = row[9]    # Column J: broj osiguranika
        district = row[13]   # Column N: grad/općina/gradska četvrt

        if not pu_name or not patients:
            continue

        pu_name = pu_name.strip()
        patients = int(patients) if patients else 0

        # Determine county ID
        if pu_name == 'PU Zagreb':
            district_upper = (district or '').strip().upper()
            if district_upper in ZAGREB_DISTRICTS:
                county_id = 'ZG'
            else:
                county_id = 'ZK'
        else:
            county_id = PU_TO_COUNTY.get(pu_name)
            if not county_id:
                print(f"  Warning: Unknown PU '{pu_name}'")
                continue

        if county_id not in gp_data:
            gp_data[county_id] = {'doctors': 0, 'patients': 0}
        gp_data[county_id]['doctors'] += 1
        gp_data[county_id]['patients'] += patients

    for cid, data in gp_data.items():
        data['per_doctor'] = round(data['patients'] / data['doctors'], 1) if data['doctors'] > 0 else 0

    print(f"GP data: {len(gp_data)} counties, {sum(d['doctors'] for d in gp_data.values())} doctors total")
    return gp_data


def classify_health_risk(patients_per_doctor):
    """Classify GP availability risk level."""
    if patients_per_doctor < 1500:
        return 'zeleno'
    elif patients_per_doctor <= 1800:
        return 'zuto'
    else:
        return 'crveno'


def compute_risk_score(pad_postotak, ucenici_pad_pct, pacijenti_po_doktoru):
    """Compute composite risk score (0-100). Higher = worse."""
    # Normalize demographic decline: -25% → 100, 0% → 0
    demo_norm = min(100, max(0, abs(pad_postotak) / 25 * 100))

    # Normalize pupil decline: -40% → 100, 0% → 0
    school_norm = min(100, max(0, abs(ucenici_pad_pct) / 40 * 100))

    # Normalize GP ratio: 2200 → 100, 1000 → 0
    health_norm = min(100, max(0, (pacijenti_po_doktoru - 1000) / 1200 * 100))

    # Weighted composite: demographic 40%, school 30%, health 30%
    score = demo_norm * 0.4 + school_norm * 0.3 + health_norm * 0.3
    return round(score, 1)


def main():
    print("=== Processing school and healthcare data ===\n")

    # Load zupanije.json for demographic decline data
    with open(ZUPANIJE_PATH, 'r', encoding='utf-8') as f:
        zupanije = json.load(f)
    zup_map = {z['id']: z for z in zupanije}

    # Read data
    school_2024 = read_school_2024()
    school_2010 = read_school_2010()
    gp = read_gp_data()

    print()

    # Build output
    results = []
    all_ids = sorted(set(list(school_2024.keys()) + list(school_2010.keys()) + list(gp.keys())))

    for cid in all_ids:
        entry = {'id': cid}

        # School data
        s2024 = school_2024.get(cid, 0)
        s2010 = school_2010.get(cid, 0)
        entry['ucenici_2024'] = s2024
        entry['ucenici_2010'] = s2010
        if s2010 > 0:
            pad_pct = round(((s2024 - s2010) / s2010) * 100, 1)
        else:
            pad_pct = 0
        entry['ucenici_pad_pct'] = pad_pct
        entry['rizik_skolski'] = pad_pct < -25  # > 25% decline

        # GP data
        gp_entry = gp.get(cid, {'doctors': 0, 'patients': 0, 'per_doctor': 0})
        entry['br_doktora'] = gp_entry['doctors']
        entry['br_pacijenata'] = gp_entry['patients']
        entry['pacijenti_po_doktoru'] = gp_entry['per_doctor']
        entry['rizik_zdravstveni'] = classify_health_risk(gp_entry['per_doctor'])

        # Composite risk
        pad_demo = zup_map[cid]['pad_postotak'] if cid in zup_map else 0
        entry['rizik_ukupni'] = compute_risk_score(pad_demo, pad_pct, gp_entry['per_doctor'])

        results.append(entry)

    # Sort by risk score descending
    results.sort(key=lambda x: x['rizik_ukupni'], reverse=True)

    # Write output
    os.makedirs(os.path.dirname(OUT_PATH), exist_ok=True)
    with open(OUT_PATH, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)

    print(f"Output: {OUT_PATH}")
    print(f"Counties: {len(results)}")
    print()

    # Summary
    school_risk = [r for r in results if r['rizik_skolski']]
    health_red = [r for r in results if r['rizik_zdravstveni'] == 'crveno']
    health_yellow = [r for r in results if r['rizik_zdravstveni'] == 'zuto']

    print("=== Summary ===")
    print(f"Counties with >25% pupil decline: {len(school_risk)}")
    for r in school_risk:
        name = zup_map[r['id']]['naziv'] if r['id'] in zup_map else r['id']
        print(f"  {name}: {r['ucenici_pad_pct']}%")

    print(f"\nCounties with GP overload (>1800 patients/doctor): {len(health_red)}")
    for r in health_red:
        name = zup_map[r['id']]['naziv'] if r['id'] in zup_map else r['id']
        print(f"  {name}: {r['pacijenti_po_doktoru']} patients/doctor")

    print(f"\nCounties with moderate GP load (1500-1800): {len(health_yellow)}")
    for r in health_yellow:
        name = zup_map[r['id']]['naziv'] if r['id'] in zup_map else r['id']
        print(f"  {name}: {r['pacijenti_po_doktoru']} patients/doctor")

    print(f"\nTop 5 composite risk:")
    for r in results[:5]:
        name = zup_map[r['id']]['naziv'] if r['id'] in zup_map else r['id']
        print(f"  {name}: {r['rizik_ukupni']} (school: {r['ucenici_pad_pct']}%, GP: {r['pacijenti_po_doktoru']})")


if __name__ == '__main__':
    main()
